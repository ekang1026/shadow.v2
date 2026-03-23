"""
Script 1 — PitchBook Ingestion
Reads PitchBook Excel/CSV export and ingests companies into Supabase.
Supports both .xlsx and .csv formats.
"""

import csv
import io
import os
import sys
import logging
from datetime import datetime
from pathlib import Path

logging.basicConfig(level=logging.INFO, format="%(asctime)s [PitchBook] %(message)s")
log = logging.getLogger(__name__)


# Maps PitchBook export column headers → snapshot database fields
# All PitchBook-sourced fields use pb_ prefix to distinguish from scraped data
COLUMN_MAP = {
    # Identity
    "Companies": "name",
    "Company ID": "pb_company_id",
    "PBId": "pitchbook_id",           # Used as the unique key on companies table
    "Company Legal Name": "pb_legal_name",
    "Description": "pb_description",

    # Location
    "HQ Location": "location",
    "HQ City": "pb_hq_city",
    "HQ State/Province": "pb_hq_state",

    # Company details
    "Year Founded": "founded_year",
    "Employees": "pb_employees",       # PitchBook headcount (unreliable — LinkedIn scrape is trusted)
    "Keywords": "pb_keywords",
    "Financing Status Note": "pb_financing_status",

    # Funding
    "Total Raised": "total_capital_raised",
    "Last Known Valuation": "last_round_valuation",
    "Last Known Valuation Date": "pb_valuation_date",
    "Last Financing Date": "pb_last_financing_date",
    "Last Financing Size": "pb_last_financing_size",
    "Active Investors": "pb_active_investors",

    # URLs
    "Website": "website",
    "LinkedIn URL": "linkedin_url",
    "View Company Online": "pb_profile_url",

    # Primary contact
    "Primary Contact": "pb_primary_contact",
    "Primary Contact Email": "pb_primary_contact_email",
    "Primary Contact Phone": "pb_primary_contact_phone",
    "Primary Contact Title": "pb_primary_contact_title",
}

# Fields that should be parsed as numbers (currency amounts)
MONEY_FIELDS = {"total_capital_raised", "last_round_valuation", "pb_last_financing_size"}

# Fields that should be parsed as integers
INT_FIELDS = {"founded_year", "pb_employees"}


def parse_number(val) -> int | float | None:
    """Parse a number from PitchBook format (e.g., '$5M', '5,000', etc.)."""
    if val is None:
        return None
    val = str(val).strip()
    if not val or val in ("", "-", "N/A", "nan", "None"):
        return None
    val = val.replace(",", "").replace("$", "").replace(" ", "")
    multiplier = 1
    if val.upper().endswith("B"):
        multiplier = 1_000_000_000
        val = val[:-1]
    elif val.upper().endswith("M"):
        multiplier = 1_000_000
        val = val[:-1]
    elif val.upper().endswith("K"):
        multiplier = 1_000
        val = val[:-1]
    try:
        result = float(val) * multiplier
        return int(result) if result == int(result) else result
    except ValueError:
        return None


def clean_value(val) -> str | None:
    """Clean a string value from PitchBook. Returns None for empty/NaN."""
    if val is None:
        return None
    val = str(val).strip()
    if val in ("", "-", "N/A", "nan", "None", "NaN"):
        return None
    return val


def parse_hyperlink(val) -> str | None:
    """Extract URL from Excel HYPERLINK formula. e.g. =HYPERLINK("http://...", "text") → http://..."""
    if val is None:
        return None
    val = str(val).strip()
    if val.startswith('=HYPERLINK('):
        # Extract first quoted string (the URL)
        import re
        match = re.search(r'=HYPERLINK\("([^"]+)"', val)
        if match:
            return match.group(1)
    return val if val and val not in ("", "-", "N/A", "nan", "None") else None


def read_excel(file_path: str) -> list[dict]:
    """
    Read a PitchBook .xlsx file. Data starts at row 9 (header row 8, 0-indexed row 7).
    Handles HYPERLINK formulas in LinkedIn URL, Website, and View Company Online columns.
    """
    from openpyxl import load_workbook

    wb = load_workbook(file_path, data_only=False)  # data_only=False to see formulas
    ws = wb.active

    # Get headers from row 8
    header_row = 8
    headers = [cell.value for cell in ws[header_row] if cell.value is not None]
    num_cols = len(headers)

    # Columns that may contain HYPERLINK formulas
    hyperlink_columns = {"LinkedIn URL", "Website", "View Company Online"}

    rows = []
    for row_num in range(header_row + 1, ws.max_row + 1):
        row_data = {}
        for col_idx in range(num_cols):
            cell = ws.cell(row=row_num, column=col_idx + 1)
            header = headers[col_idx]
            val = cell.value

            # Parse HYPERLINK formulas for URL columns
            if header in hyperlink_columns and val and str(val).startswith("=HYPERLINK"):
                row_data[header] = parse_hyperlink(val)
            else:
                row_data[header] = val

        # Skip empty rows
        if row_data.get("Companies") or row_data.get("Company ID"):
            rows.append(row_data)

    log.info(f"Read {len(rows)} rows from Excel (parsed HYPERLINK formulas)")
    return rows


def read_csv_content(csv_content: str) -> list[dict]:
    """Read CSV content string into list of dicts."""
    reader = csv.DictReader(io.StringIO(csv_content))
    return list(reader)


def build_snapshot_data(row: dict, pitchbook_id: str) -> dict | None:
    """Build snapshot data dict from a PitchBook row."""
    snapshot_data = {}
    for csv_col, field in COLUMN_MAP.items():
        if csv_col not in row or field == "pitchbook_id":
            continue
        raw_val = row[csv_col]
        if field in MONEY_FIELDS:
            parsed = parse_number(raw_val)
            snapshot_data[field] = int(parsed) if parsed is not None else None
        elif field in INT_FIELDS:
            parsed = parse_number(raw_val)
            snapshot_data[field] = int(parsed) if parsed is not None else None
        else:
            cleaned = clean_value(raw_val)
            if cleaned:
                snapshot_data[field] = cleaned

    if snapshot_data.get("pb_profile_url") == "View Company Online":
        snapshot_data["pb_profile_url"] = f"https://pitchbook.com/profiles/company/{pitchbook_id}"
    snapshot_data["pb_company_id"] = pitchbook_id
    return snapshot_data


def ingest_rows(rows: list[dict]) -> dict:
    """
    Process rows and upsert companies + snapshots into Supabase.
    Uses batch operations for speed (~10x faster than row-by-row).
    Returns stats: {new, updated, skipped, errors}
    """
    from config import get_supabase

    sb = get_supabase()
    stats = {"new": 0, "updated": 0, "skipped": 0, "errors": 0}
    total = len(rows)

    # Step 1: Parse all rows to get pitchbook IDs
    log.info(f"Parsing {total} rows...")
    parsed_rows = []
    for row in rows:
        pitchbook_id = clean_value(row.get("PBId")) or clean_value(row.get("Company ID"))
        if not pitchbook_id:
            company_name = clean_value(row.get("Companies"))
            if company_name:
                pitchbook_id = f"PB-{company_name[:50]}"
            else:
                stats["skipped"] += 1
                continue
        parsed_rows.append((pitchbook_id, row))

    # Step 2: Fetch all existing companies in one call
    log.info(f"Checking {len(parsed_rows)} companies against database...")
    all_pb_ids = [pb_id for pb_id, _ in parsed_rows]

    # Fetch in batches of 500 (Supabase URL length limit)
    existing_companies = {}
    for i in range(0, len(all_pb_ids), 500):
        batch_ids = all_pb_ids[i:i+500]
        result = sb.table("companies").select("id,pitchbook_id,status").in_("pitchbook_id", batch_ids).execute()
        for c in result.data:
            existing_companies[c["pitchbook_id"]] = c

    log.info(f"Found {len(existing_companies)} existing companies")

    # Step 3: Batch insert new companies
    new_company_ids = [pb_id for pb_id in all_pb_ids if pb_id not in existing_companies]
    if new_company_ids:
        # Deduplicate
        unique_new = list(set(new_company_ids))
        log.info(f"Creating {len(unique_new)} new companies...")
        for i in range(0, len(unique_new), 200):
            batch = [{"pitchbook_id": pb_id, "status": "pending", "review_count": 0} for pb_id in unique_new[i:i+200]]
            result = sb.table("companies").insert(batch).execute()
            for c in result.data:
                existing_companies[c["pitchbook_id"]] = c

    # Step 4: Process snapshots in batches
    log.info(f"Processing snapshots...")
    snapshot_batch = []
    processed = 0

    for pitchbook_id, row in parsed_rows:
        try:
            company = existing_companies.get(pitchbook_id)
            if not company:
                stats["errors"] += 1
                continue

            # Skip already classified
            if company["status"] != "pending":
                stats["skipped"] += 1
                continue

            snapshot_data = build_snapshot_data(row, pitchbook_id)
            if not snapshot_data:
                stats["skipped"] += 1
                continue

            snapshot_data["company_id"] = company["id"]
            snapshot_data["is_latest"] = True
            snapshot_data["snapshot_date"] = datetime.now().date().isoformat()
            # Explicitly set filters to None — not yet evaluated
            # (DB defaults to false, but we want null = "pending evaluation")
            snapshot_data["passed_headcount_filter"] = None
            snapshot_data["passed_llm_filter"] = None
            snapshot_batch.append(snapshot_data)
            stats["new"] += 1
            processed += 1

            # Log progress every 100 rows
            if processed % 100 == 0:
                log.info(f"Processed {processed}/{total} rows...")

        except Exception as e:
            log.error(f"Error processing row {pitchbook_id}: {e}")
            stats["errors"] += 1

    # Step 5: Mark old snapshots as not latest (batch by company_ids)
    if snapshot_batch:
        company_ids_to_update = list(set(s["company_id"] for s in snapshot_batch))
        log.info(f"Marking {len(company_ids_to_update)} old snapshots as not latest...")
        for i in range(0, len(company_ids_to_update), 500):
            batch_ids = company_ids_to_update[i:i+500]
            sb.table("company_snapshots") \
                .update({"is_latest": False}) \
                .in_("company_id", batch_ids) \
                .eq("is_latest", True) \
                .execute()

    # Step 6: Batch insert new snapshots
    if snapshot_batch:
        log.info(f"Inserting {len(snapshot_batch)} snapshots...")
        for i in range(0, len(snapshot_batch), 200):
            batch = snapshot_batch[i:i+200]
            try:
                sb.table("company_snapshots").insert(batch).execute()
                log.info(f"Inserted batch {i//200 + 1} ({len(batch)} snapshots)")
            except Exception as e:
                log.error(f"Batch insert error: {e}")
                # Fall back to row-by-row for this batch
                for snap in batch:
                    try:
                        sb.table("company_snapshots").insert(snap).execute()
                    except Exception as e2:
                        log.error(f"Row insert error for {snap.get('name', '?')}: {e2}")
                        stats["errors"] += 1
                        stats["new"] -= 1

    return stats


def download_csvs_from_pitchbook() -> list[str]:
    """
    Use Playwright to navigate to PitchBook, log in, and export CSVs from all saved searches.
    Returns list of CSV content strings.

    IMPORTANT: Do NOT call this unless the user explicitly requests a PitchBook download.
    PitchBook has a 10K monthly download limit across both saved searches.
    """
    from playwright.sync_api import sync_playwright
    from config import PITCHBOOK_EMAIL, PITCHBOOK_PASSWORD, PITCHBOOK_SEARCH_URLS

    if not PITCHBOOK_SEARCH_URLS:
        raise ValueError("No PITCHBOOK_SEARCH_URL(s) set in pipeline/.env")
    if not PITCHBOOK_EMAIL or not PITCHBOOK_PASSWORD:
        raise ValueError("PITCHBOOK_EMAIL and PITCHBOOK_PASSWORD must be set in pipeline/.env")

    log.info(f"Launching browser for PitchBook export ({len(PITCHBOOK_SEARCH_URLS)} searches)...")

    csv_contents = []

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False)
        context = browser.new_context()
        page = context.new_page()

        page.goto("https://pitchbook.com/login")
        page.wait_for_load_state("networkidle")

        page.fill('input[name="email"], input[type="email"]', PITCHBOOK_EMAIL)
        page.fill('input[name="password"], input[type="password"]', PITCHBOOK_PASSWORD)
        page.click('button[type="submit"]')
        page.wait_for_load_state("networkidle")
        log.info("Logged into PitchBook")

        download_path = os.path.join(os.path.dirname(__file__), "downloads")
        os.makedirs(download_path, exist_ok=True)

        for i, search_url in enumerate(PITCHBOOK_SEARCH_URLS):
            log.info(f"Navigating to saved search {i + 1}/{len(PITCHBOOK_SEARCH_URLS)}: {search_url}")
            page.goto(search_url)
            page.wait_for_load_state("networkidle")

            with page.expect_download() as download_info:
                export_btn = page.locator('button:has-text("Export"), button:has-text("Download"), [data-testid="export"]')
                if export_btn.count() > 0:
                    export_btn.first.click()
                else:
                    page.click('button:has-text("Actions"), button:has-text("More")')
                    page.click('text=Export to CSV, text=Download CSV, text=Export')

            download = download_info.value
            csv_path = os.path.join(download_path, download.suggested_filename)
            download.save_as(csv_path)
            log.info(f"Downloaded CSV {i + 1}: {csv_path}")

            with open(csv_path, "r", encoding="utf-8") as f:
                csv_contents.append(f.read())

        browser.close()

    log.info(f"Downloaded {len(csv_contents)} CSV files from PitchBook")
    return csv_contents


def run(file_path: str = None, pitchbook_download: bool = False) -> dict:
    """
    Main entry point.
    - file_path: local .xlsx or .csv file to ingest
    - pitchbook_download: set to True ONLY when user explicitly requests PitchBook download

    IMPORTANT: Never set pitchbook_download=True automatically.
    PitchBook has a 10K monthly download limit.
    """
    all_rows = []

    if file_path:
        ext = Path(file_path).suffix.lower()
        log.info(f"Reading from local file: {file_path}")

        if ext in (".xlsx", ".xls"):
            all_rows = read_excel(file_path)
        elif ext == ".csv":
            with open(file_path, "r", encoding="utf-8") as f:
                all_rows = read_csv_content(f.read())
        else:
            raise ValueError(f"Unsupported file format: {ext}. Use .xlsx or .csv")

    elif pitchbook_download:
        log.info("PITCHBOOK DOWNLOAD: User explicitly requested download from PitchBook")
        csv_contents = download_csvs_from_pitchbook()
        for csv_content in csv_contents:
            all_rows.extend(read_csv_content(csv_content))
    else:
        raise ValueError(
            "No file provided and pitchbook_download not enabled. "
            "Pass a file path or set pitchbook_download=True (only when user explicitly requests it)."
        )

    log.info(f"Found {len(all_rows)} rows to process")
    stats = ingest_rows(all_rows)

    log.info(f"PitchBook ingestion complete: {stats['new']} new, {stats['updated']} updated, "
             f"{stats['skipped']} skipped, {stats['errors']} errors")
    return stats


if __name__ == "__main__":
    if len(sys.argv) > 1:
        arg = sys.argv[1]
        if arg == "--pitchbook":
            run(pitchbook_download=True)
        else:
            run(file_path=arg)
    else:
        print("Usage:")
        print("  python3 script1_pitchbook.py /path/to/export.xlsx   # Ingest local file")
        print("  python3 script1_pitchbook.py --pitchbook            # Download from PitchBook (uses monthly quota!)")
        sys.exit(1)
